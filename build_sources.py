import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "크롤링 대상 소스"

FONT_NAME = "Arial"
headers = ["계층", "소스명", "URL", "유형", "수집 방식", "제공 데이터", "난이도", "활용 방안 / 비고", "진행상태"]

rows = [
    # Tier 1: 메이저 직접소스 (수동/개별 확인, 정확도 최우선)
    ["Tier1", "서울마라톤(동아마라톤)", "https://seoul-marathon.com/", "메이저 대회 공식",
     "직접 확인(수동)", "정확한 접수기간 / 정원 / 코스 / 공지사항", "낮음 (공식 사이트, 구조 명확)",
     "국내 최상위 메이저. 정확도가 핵심이므로 담당자가 주기적으로 직접 확인", "예정"],
    ["Tier1", "JTBC 서울마라톤", "http://marathon.jtbc.com/", "메이저 대회 공식",
     "직접 확인(수동)", "정확한 접수기간 / 추첨 결과 / 공지사항", "낮음",
     "추첨제 구조라 접수 일정 변동이 잦음 - 공지 확인 주기를 짧게 가져가야 함", "예정"],
    ["Tier1", "춘천마라톤", "https://www.chuncheonmarathon.com/", "메이저 대회 공식",
     "직접 확인(수동)", "정확한 접수기간(선착순) / 정원 / 코스", "낮음",
     "선착순 접수라 마감 시점 정확도가 특히 중요", "예정"],
    ["Tier1", "보스턴 마라톤 (BAA)", "https://www.baa.org/races/boston-marathon/", "해외 메이저 공식",
     "직접 확인(수동)", "접수기간 / 자격기준(BQ타임) / 공지사항", "낮음",
     "해외 메이저는 국내 통합 사이트가 거의 다루지 않아 100% 직접 소스 필요", "예정"],
    ["Tier1", "도쿄 마라톤", "https://www.marathon.tokyo/en/", "해외 메이저 공식",
     "직접 확인(수동)", "접수기간(추첨) / 공지사항", "낮음",
     "추첨제. 접수 오픈·결과 발표 일정 확인 필요", "예정"],
    ["Tier1", "TCS 뉴욕시티 마라톤", "https://www.nyrr.org/tcsnycmarathon", "해외 메이저 공식",
     "직접 확인(수동)", "접수기간(추첨) / 공지사항", "낮음",
     "NYRR 공식 서브도메인이 실질적 공식 페이지", "예정"],
    ["Tier1", "시카고 마라톤", "https://www.chicagomarathon.com/", "해외 메이저 공식",
     "직접 확인(수동)", "접수기간(추첨) / 공지사항", "낮음",
     "월드마라톤메이저스 소속 - 국내 러너 관심도 높음", "예정"],

    # Tier 2: 국내 롱테일 - 통합 사이트 자동 크롤링
    ["Tier2", "고러닝", "https://gorunning.kr/races/", "통합 정보사이트(국내 롱테일)",
     "자동 크롤링", "대회명 / 일정 / 지역 / 거리 / 접수상태 / 접수기간", "낮음 (정적 HTML, 구조 명확)",
     "국내 롱테일 커버리지의 기본 축. 상태기반 UX도 함께 벤치마킹", "예정"],
    ["Tier2", "마라톤GO", "https://marathongo.co.kr/raceSchedule/domestic", "통합 정보사이트(국내 롱테일)",
     "자동 크롤링", "대회명 / 일정 / 지역 / 거리 / 접수상태 / 장소 / 집결시간 / 주최", "중간 (Next.js SPA, API 확인 필요)",
     "필터 항목이 가장 상세. 대회 상세정보 보강용", "예정"],
    ["Tier2", "러닝위키", "https://runningwikii.com/", "통합 정보사이트(국내 롱테일)",
     "자동 크롤링", "대회명 / 일정 / 지역 / 거리 (월별 표)", "낮음 (정적 HTML 표)",
     "신규 대회 발굴 및 교차검증용. 접수상태는 부실해 참고용", "예정"],
    ["Tier2", "로드런", "http://www.roadrun.co.kr/schedule/list.php", "통합 정보사이트(레거시, 국내 롱테일)",
     "자동 크롤링", "대회명 / 일정 / 장소 / 주최 / 연락처", "낮음 (구형 정적 HTML)",
     "등록 대회 수(437개+)가 많아 롱테일 보강에 유리", "예정"],
    ["Tier2", "전국마라톤협회", "http://mara1080.com/", "협회 공식(국내 롱테일)",
     "자동 크롤링 시도 → 실패 시 수동", "대회 등록 / 신청 / 결과", "높음 (JS 렌더링, 정적 크롤링 콘텐츠 미노출)",
     "협회 승인 대회 확인용. Playwright 등 브라우저 렌더링 필요", "예정"],
    ["Tier2", "러너블", "https://runable.me/competition", "접수대행 플랫폼(교차검증)",
     "자동 크롤링", "대회명 / 접수기간 / 거리 / 접수상태", "중간 (SPA)",
     "실제 접수 링크·마감일 교차검증용", "예정"],
    ["Tier2", "랭킹마라톤", "https://rankingmarathon.com/", "접수대행 플랫폼(교차검증)",
     "자동 크롤링", "대회명 / 국가 / 연도 / 월 / 거리 필터", "중간 (SPA)",
     "해외 대회 검색 기능 보유 - 해외 큐레이션 보강 소스로도 활용", "예정"],

    # Tier 3: 발굴/보강용
    ["Tier3", "Ahotu", "https://www.ahotu.com/ko/calendar/running/marathon", "글로벌 통합 DB(해외 발굴)",
     "자동 크롤링", "전세계 대회명 / 일정 / 국가 / 거리 / 가격", "중간 (다국어 SPA)",
     "Tier1 외 해외 대회(15%) 발굴용. 발견 후 공식 사이트로 검증", "예정"],
    ["Tier3", "FindMyMarathon", "https://findmymarathon.com/calendar.php", "데이터중심 큐레이션(해외 발굴)",
     "자동 크롤링", "코스 고도 / 난이도 / PR스코어 / BQ 여부", "낮음 (정적 HTML 리스트)",
     "해외 대회 발굴 + 3단계 '의사결정 지원 데이터' 설계 참고 모델", "예정"],
    ["Tier3", "지자체 체육회 / 육상연맹 홈페이지", "다수 개별 URL", "지역 소규모 대회 주관처(국내 발굴)",
     "수동 확인", "지역 소규모 대회 공지", "높음 (사이트마다 형식 상이, 비정형)",
     "롱테일 중에서도 가장 파편화된 영역. 자동화는 3단계 이후 확장 검토", "예정"],
]

col_widths = [8, 24, 40, 26, 18, 34, 30, 44, 10]

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
tier_fill = {
    "Tier1": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
    "Tier2": PatternFill(start_color="FDF2E9", end_color="FDF2E9", fill_type="solid"),
    "Tier3": PatternFill(start_color="EAF7EA", end_color="EAF7EA", fill_type="solid"),
}

for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = col_widths[c - 1]

for r, row in enumerate(rows, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT_NAME, size=10.5)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        if c == 1:
            cell.fill = tier_fill.get(val, PatternFill())
            cell.font = Font(name=FONT_NAME, size=10.5, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 44

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# Legend sheet
legend = wb.create_sheet("범례")
legend_font_bold = Font(name=FONT_NAME, bold=True, size=11)
legend_font = Font(name=FONT_NAME, size=10.5)
legend.column_dimensions["A"].width = 10
legend.column_dimensions["B"].width = 78
legend["A1"] = "계층"
legend["B1"] = "설명"
legend["A1"].font = legend_font_bold
legend["B1"].font = legend_font_bold
legend_rows = [
    ("Tier1", "메이저/시그니처 대회(20~30개 내외). 정확도가 핵심이라 개별 공식 사이트를 직접 확인/관리한다. 검색 트래픽과 정확도 요구가 가장 높은 영역."),
    ("Tier2", "국내 롱테일 대회(수백 개). 고러닝·마라톤GO 등 통합 사이트를 자동 크롤링해 커버리지를 빠르게 확보한다. 정확도는 100%를 기대하지 않고 공식 링크로 보완."),
    ("Tier3", "발굴/보강용. 해외 대회 탐색(Ahotu, FindMyMarathon)과 지역 소규모 대회 발굴(지자체 체육회)에 활용하며, 자동화 난이도가 높아 확장 단계에서 다룬다."),
    ("진행상태", "예정 / 진행중 / 완료 로 직접 갱신하며 트래킹."),
]
for i, (a, b) in enumerate(legend_rows, start=2):
    legend.cell(row=i, column=1, value=a).font = legend_font
    legend.cell(row=i, column=2, value=b).font = legend_font
    legend.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
    legend.row_dimensions[i].height = 40

# ---------- Tier1 실데이터 (races 테이블 시범 채우기) ----------
t1data = wb.create_sheet("Tier1 실데이터")
t1_headers = ["대회명", "대회일", "접수상태(2026.7.27 기준)", "접수기간", "모집인원/정원", "방식", "공식링크", "last_verified_at", "비고"]
t1_col_widths = [22, 14, 20, 30, 16, 14, 34, 16, 40]

t1_rows = [
    ["서울마라톤(동아마라톤)", "2027.03.21", "접수전",
     "기록제출 4.28~5.11 / 우선접수 5.18~22 / 본접수 풀 6.1~2, 10km 6.3",
     "미공개", "그룹별 상이", "https://seoul-marathon.com/", "2026-07-27",
     "2027년 대회. 접수는 내년 4~6월 예정"],
    ["JTBC 서울마라톤", "2026.11.01", "접수마감",
     "접수 4.13 14시~4.15 16시 / 당첨발표 4.17", "풀 약 1만 / 10km 약 1만",
     "기록기반 추첨(래플)", "http://marathon.jtbc.com/", "2026-07-27",
     "본접수는 4월에 이미 종료됨"],
    ["춘천마라톤", "2026.10.25", "접수마감",
     "일반 Full 신청 7.14~ 결제 7.20~21 / 10km 신청 7.16~ 결제 7.22~23",
     "미공개", "선착순", "https://www.chuncheonmarathon.com/", "2026-07-27",
     "전 부문 결제기간이 이미 지남(7.23 마감)"],
    ["보스턴 마라톤(BAA)", "2027.04.19", "접수전",
     "자격자(Qualifier) 등록 9.14~9.18(2026)", "약 30,000명",
     "BQ기록 + 추가 랜덤선발", "https://www.baa.org/races/boston-marathon/", "2026-07-27",
     "131회 대회. 등록일정은 7.22 공식 발표됨"],
    ["도쿄 마라톤", "2027.03.07", "부문별 상이",
     "채리티 6.24~7.9(마감) / ONE TOKYO·준엘리트 7.31~8.13(예정) / 일반추첨 8.14~8.28(예정)",
     "일반추첨 약 3만 명(경쟁률 약 10:1)", "추첨(부문별)", "https://www.marathon.tokyo/en/", "2026-07-27",
     "부문마다 접수 시기가 달라 상태를 하나로 못 묶음 - UI 설계 시 주의"],
    ["TCS 뉴욕시티 마라톤", "2026.11.01", "접수마감",
     "일반추첨 신청 2.4~2.25 / 추첨발표 3.4(2026년 기준 이미 종료)",
     "24만명+ 응모(역대 최다)", "추첨(지역별 3개 풀)", "https://www.nyrr.org/tcsnycmarathon", "2026-07-27",
     "공식 사이트 첫 화면 자체가 2024/2025년 캐시 데이터를 보여줌 - 원문도 100% 신뢰 금지, 교차검증 필요"],
    ["시카고 마라톤", "2026.10.11", "접수마감",
     "일반 신청 접수 종료(공식 공지) - 이후 투어프로그램 통해서만 참가 가능",
     "미공개", "추첨 + 투어프로그램", "https://www.chicagomarathon.com/", "2026-07-27",
     "일반 접수창은 닫혔고 Tour Program 경로만 남음"],
]

header_fill2 = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
status_fill = {
    "접수전": PatternFill(start_color="EAF7EA", end_color="EAF7EA", fill_type="solid"),
    "접수마감": PatternFill(start_color="FDEBEB", end_color="FDEBEB", fill_type="solid"),
    "부문별 상이": PatternFill(start_color="FDF2E9", end_color="FDF2E9", fill_type="solid"),
}

for c, h in enumerate(t1_headers, start=1):
    cell = t1data.cell(row=1, column=c, value=h)
    cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill2
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    t1data.column_dimensions[get_column_letter(c)].width = t1_col_widths[c - 1]

for r, row in enumerate(t1_rows, start=2):
    for c, val in enumerate(row, start=1):
        cell = t1data.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT_NAME, size=10.5)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        if c == 3:
            cell.fill = status_fill.get(val, PatternFill())
            cell.font = Font(name=FONT_NAME, size=10.5, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    t1data.row_dimensions[r].height = 56

t1data.freeze_panes = "A2"
t1data.auto_filter.ref = f"A1:{get_column_letter(len(t1_headers))}{len(t1_rows)+1}"

wb.save("크롤링_대상_소스_리스트.xlsx")
print("saved")
